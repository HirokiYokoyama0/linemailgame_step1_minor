import openpyxl
import random

def create_word(GENRENUM):

  workbook = openpyxl.load_workbook('LINEマイノリティゲーム_データベース_No1.xlsx')

  #sheet = workbook["General"]
  sheet=workbook.worksheets[GENRENUM-1]
  print("sheet名",sheet)
  max_row_num = sheet.max_row
  LINEqa_data = []

  for i in range(2,max_row_num+1):
    mailmain = sheet.cell(row=i, column=1).value #メール文例
    ans1 = sheet.cell(row=i, column=2).value #返答例1
    ans2 = sheet.cell(row=i, column=3).value #返答例2
    ans3 = sheet.cell(row=i, column=4).value #返答例3
    ans4 = sheet.cell(row=i, column=5).value #返答例4
    ans5 = sheet.cell(row=i, column=6).value #返答例5

    #if cell_value not in suppliers:
    LINEqa_data.append([mailmain,ans1,ans2,ans3,ans4,ans5])

  #print("max_row_num==>",max_row_num,len(word_data))
  word_num = random.randint(0,len(LINEqa_data)-1)
  #print("word_num-->",word_num)
  #print("word_data[0]-->",word_data[word_num])
  #print("word_data[7][0]-->",word_data[word_num][0])
  #print("word_data[7][1]-->",word_data[word_num][1])
  # 戻り地として、重複のない乱数列とデータ一覧を戻す。
  return LINEqa_data,max_row_num
  


# 重複のない乱数を生成する
def rand_ints_nodup(a, b, k):
  ns = []
  while len(ns) < k:
    n = random.randint(a, b)
    if not n in ns:
      ns.append(n)
  return ns


def check_genre():
  workbook = openpyxl.load_workbook('LINEマイノリティゲーム_データベース_No1.xlsx')
  sheets = workbook.sheetnames

  print(sheets)

  return sheets



